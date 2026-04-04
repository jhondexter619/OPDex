"""Excel file operations: create, write, append, read, format."""

import argparse
import json
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils import output_json, setup_logging, timestamp

log = setup_logging("excel_ops")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_workbook(
    file_path: str,
    sheets: dict[str, list[list]] | None = None,
    auto_format: bool = True,
) -> dict:
    """Create a new Excel workbook with optional sheet data.

    Args:
        file_path: Path to save the .xlsx file.
        sheets: Dict of {sheet_name: [[header_row], [data_row], ...]}.
        auto_format: Apply header formatting and auto-column-width.
    """
    log.info("Creating workbook: %s", file_path)

    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet if we have named sheets
    if sheets:
        wb.remove(wb.active)

    if sheets:
        for sheet_name, rows in sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            for row in rows:
                ws.append(row)
            if auto_format and rows:
                _format_header(ws)
                _auto_width(ws)
    else:
        ws = wb.active
        ws.title = "Sheet1"

    wb.save(file_path)
    log.info("Saved workbook with %d sheets", len(wb.sheetnames))

    return {
        "success": True,
        "file_path": os.path.abspath(file_path),
        "sheets": wb.sheetnames,
        "timestamp": timestamp(),
    }


def append_rows(
    file_path: str,
    rows: list[list],
    sheet_name: str = "Sheet1",
) -> dict:
    """Append rows to an existing worksheet."""
    log.info("Appending %d rows to %s in %s", len(rows), sheet_name, file_path)

    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)

    ws = wb[sheet_name]
    for row in rows:
        ws.append(row)

    _auto_width(ws)
    wb.save(file_path)

    return {
        "success": True,
        "file_path": os.path.abspath(file_path),
        "sheet": sheet_name,
        "rows_appended": len(rows),
        "timestamp": timestamp(),
    }


def read_workbook(file_path: str, sheet_name: str | None = None) -> dict:
    """Read data from an Excel workbook."""
    log.info("Reading: %s", file_path)

    wb = openpyxl.load_workbook(file_path, read_only=True)
    result_sheets = {}

    sheets_to_read = [sheet_name] if sheet_name else wb.sheetnames
    for name in sheets_to_read:
        ws = wb[name]
        result_sheets[name] = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]

    return {
        "success": True,
        "file_path": os.path.abspath(file_path),
        "sheets": result_sheets,
        "timestamp": timestamp(),
    }


def _format_header(ws) -> None:
    """Apply formatting to the first row."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def _auto_width(ws) -> None:
    """Auto-fit column widths based on content."""
    for col_idx, col_cells in enumerate(ws.iter_cols(values_only=True), 1):
        max_len = 0
        for val in col_cells:
            if val is not None:
                max_len = max(max_len, len(str(val)))
        adjusted = min(max_len + 4, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excel file operations")
    sub = parser.add_subparsers(dest="operation", required=True)

    # create
    p_create = sub.add_parser("create")
    p_create.add_argument("file_path", help="Path to .xlsx file")
    p_create.add_argument("--sheets", type=json.loads, default=None,
                          help='JSON dict: {"SheetName": [["h1","h2"],["r1","r2"]]}')
    p_create.add_argument("--no-format", action="store_true")

    # append
    p_append = sub.add_parser("append")
    p_append.add_argument("file_path", help="Path to .xlsx file")
    p_append.add_argument("--rows", type=json.loads, required=True)
    p_append.add_argument("--sheet", default="Sheet1")

    # read
    p_read = sub.add_parser("read")
    p_read.add_argument("file_path", help="Path to .xlsx file")
    p_read.add_argument("--sheet", default=None)

    args = parser.parse_args()

    if args.operation == "create":
        result = create_workbook(args.file_path, args.sheets, not args.no_format)
    elif args.operation == "append":
        result = append_rows(args.file_path, args.rows, args.sheet)
    elif args.operation == "read":
        result = read_workbook(args.file_path, args.sheet)

    output_json(result)
